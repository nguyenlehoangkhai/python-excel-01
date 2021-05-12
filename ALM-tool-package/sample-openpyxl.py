import pandas as pd
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
# Import relevant modules from `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string

from itertools import islice

# Import `dataframe_to_rows`
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *

print('######################################################################')
# Load in the workbook
wb = load_workbook('example02.xlsx')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Sheet1']
print(sheet)

# Print the sheet title
print('Sheet Title:', sheet.title)

# Get currently active sheet
anotherSheet = wb.active

# Check `anotherSheet`
print(anotherSheet)

# Retrieve the value of a certain cell
print(sheet['A1'].value)

# Select element 'B3' of your sheet
c = sheet['B3']

# Retrieve the row number of your element
print('Row No.:', c.row)

# Retrieve the column number of your element
print('Column Letter:', c.column)

# Retrieve the coordinates of the cell
print('Coordinates of cell:', c.coordinate)

# Retrieve cell value
print(sheet.cell(row=1, column=2).value)


print('########################################################################')
# Print out values in column 2
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)

# Return 'A'
print('Column Letter:', get_column_letter(1))

# Return '1'
print('Column Index:', column_index_from_string('A'))

# Print row per row
for cellObj in sheet['A1':'C3']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')

# Retrieve the maximum amount of rows
print('Max Rows:', sheet.max_row)

# Retrieve the maximum amount of columns
print('Max Columns:', sheet.max_column)

print('########################################################################')
##################################################################################

# Convert Sheet to DataFrame
dataframe = pd.DataFrame(sheet.values)
print(dataframe)

################################################################################
print('########################################################################')

# Put the sheet values in `data`
data = sheet.values

# Indicate the columns in the sheet values
cols = next(data)[1:]

# Convert your data to a list
data = list(data)

# Read in the data at index 0 for the indices
idx = [r[0] for r in data]

# Slice the data at index 1
data = (islice(r, 1, None) for r in data)

# Make your DataFrame
df = pd.DataFrame(data, index=idx, columns=cols)
print(df)


print('########################################################################')
##########################################################################

# Initialize a workbook
wb = Workbook()

# Get the worksheet in the active workbook
ws = wb.active

# Append the rows of the DataFrame to your worksheet
for r in dataframe_to_rows(dataframe, index=False, header=False):
    ws.append(r)

wb.save('new.xlsx')

######################################################################