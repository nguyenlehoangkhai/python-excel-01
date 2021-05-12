# Import `pandas`
import pandas as pd
from pandas import ExcelFile

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from itertools import islice

# Import `dataframe_to_rows`
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *

from xlwt import Workbook

wb = load_workbook('ALM_xxxx_CodeReview_ShortDescription.xlsx')
file = "ALM_xxxx_CodeReview_ShortDescription.xlsx"
path = "ALM_xxxx_CodeReview_ShortDescription.xlsx"


# Read excel file using openpyxl
'''
# Get sheet names
print(wb.sheetnames)
# Get a sheet by name
sheet = wb['Findings']
# Print the sheet title
print('Sheet Title:', sheet.title)
# Get currently active sheet
anotherSheet = wb.active
'''

# Read excel file using Pandas

xl = pd.ExcelFile(file)
print(xl.sheet_names)
# Load a sheet into a DataFrame by name: df1
df1 = xl.parse('Findings')
print(df1)
sheet4 = xl.ExcelWriter('hello')
print(xl.sheet_names)

'''
excelDf = pd.read_excel(path)
findingsSheet = excelDf.parse("Findings")
print(excelDf)
print(findingsSheet)

'''