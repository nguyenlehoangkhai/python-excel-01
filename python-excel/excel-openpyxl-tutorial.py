# Import `os`, 'panda'
import os
import pandas as pd

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


# Load in the workbook
wb = load_workbook('demo.xlsx')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Sheet1']

# Print the sheet title
print('Sheet Title:',sheet.title)

# Get currently active sheet
anotherSheet = wb.active

# Check `anotherSheet`
print(anotherSheet)

# Retrieve the value of a certain cell
print(sheet['A1'].value)

# Select element 'B3' of your sheet
c = sheet['B3']

# Retrieve the row number of your element
print('Row No.:', c.row)

# Retrieve the column number of your element
print('Column Letter:', c.column)

# Retrieve the coordinates of the cell
print('Coordinates of cell:', c.coordinate)

# Retrieve cell value
print(sheet.cell(row=1, column=2).value)

# Print out values in column 2
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)

# Import relevant modules from `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string

# Return 'A'
print('Column Letter:', get_column_letter(2))

# Return '1'
print('Column Index:', column_index_from_string('C'))


# Print row per row
for cellObj in sheet['A1':'B8']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')


# Retrieve the maximum amount of rows
print('Max Rows:', sheet.max_row)

# Retrieve the maximum amount of columns
print('Max Columns:', sheet.max_column)