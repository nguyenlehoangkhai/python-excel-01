# Import `pandas`
import pandas as pd

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from itertools import islice

# Import `dataframe_to_rows`
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *


'''
# file = 'demo.xlsx'
# # Load spreadsheet
# xl = pd.ExcelFile(file)
# # Print the sheet names
# print(xl.sheet_names)
# # Load a sheet into a DataFrame by name: df1
# df1 = xl.parse('Sheet1')
# print(df1) 
# df1.to_excel("demo_convert.xlsx")
'''


# Load in the workbook
wb = load_workbook('demo.xlsx')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Sheet1']


# Convert Sheet to DataFrame
df = pd.DataFrame(sheet.values)


# Put the sheet values in `data`
data = sheet.values

# Indicate the columns in the sheet values
cols = next(data)[1:]

# Convert your data to a list
data = list(data)

# Read in the data at index 0 for the indices
idx = [r[0] for r in data]

# Slice the data at index 1
data = (islice(r, 1, None) for r in data)

# Make your DataFrame
df = pd.DataFrame(data, index=idx, columns=cols)
print(df)

# Initialize a workbook
wb = Workbook()

# Get the worksheet in the active workbook
ws = wb.active

# Append the rows of the DataFrame to your worksheet
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

df.to_excel('demo2.xlsx')


# Check the first entries of the DataFrame
print(df.head())

# Check the last entries of the DataFrame
print(df.tail())

# Inspect the shape
print(df.shape)

# Inspect the number of dimensions
print(df.ndim)

# Inspect the data type
print(df.dtypes)